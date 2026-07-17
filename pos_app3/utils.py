import io
import csv
from datetime import datetime
from functools import wraps

from bson import ObjectId
from bson.errors import InvalidId
from flask import session, redirect, url_for, jsonify, request


# ---------- JSON helpers ----------

def to_json_safe(value):
    """Recursively convert Mongo documents (ObjectId / datetime) into JSON-safe values."""
    if isinstance(value, list):
        return [to_json_safe(v) for v in value]
    if isinstance(value, dict):
        out = {}
        for k, v in value.items():
            key = "id" if k == "_id" else k
            out[key] = to_json_safe(v)
        return out
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def parse_object_id(id_str):
    try:
        return ObjectId(id_str)
    except (InvalidId, TypeError):
        return None


# ---------- Auth ----------

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get("role") not in roles:
                return jsonify({"error": "forbidden"}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ---------- Export helpers ----------

def rows_to_csv_bytes(headers, rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return io.BytesIO(buf.getvalue().encode("utf-8-sig"))


def rows_to_xlsx_bytes(headers, rows, sheet_title="Report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for i, _ in enumerate(headers, start=1):
        max_len = max(
            [len(str(headers[i - 1]))] + [len(str(r[i - 1])) for r in rows] if rows else [len(str(headers[i - 1]))]
        )
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = min(max(12, max_len + 2), 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def invoice_number(seq: int) -> str:
    return f"INV-{datetime.now().strftime('%Y%m')}-{seq:05d}"
